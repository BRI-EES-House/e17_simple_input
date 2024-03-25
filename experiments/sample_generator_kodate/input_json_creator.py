import openpyxl


def convert_to_input_json(
    input_xlsx_filepath: str,
):
    book = openpyxl.load_workbook(input_xlsx_filepath)

    sheet_common = book['common']
    sheet_building = book['building']
    sheet_rooms = book['rooms']
    sheet_external_general_parts = book['external_general_parts']
    sheet_external_opaque_parts = book['external_opaque_parts']
    sheet_external_transparent_parts = book['external_transparent_parts']
    sheet_internals = book['internals']
    sheet_grounds = book['grounds']
    sheet_layers = book['layers']

    n_rooms = count_number_in_id_row(sheet=sheet_rooms)
    n_external_general_parts = count_number_in_id_row(sheet=sheet_external_general_parts)
    n_external_opaque_parts = count_number_in_id_row(sheet=sheet_external_opaque_parts)
    n_external_transparent_parts = count_number_in_id_row(sheet=sheet_external_transparent_parts)
    n_internals = count_number_in_id_row(sheet=sheet_internals)
    n_grounds = count_number_in_id_row(sheet=sheet_grounds)
    n_layers = count_number_in_id_row(sheet=sheet_layers)

    common = {
        'ac_method': sheet_common.cell(column=2, row=2).value
    }

    building = {
        "infiltration": {
            "method": "balance_residential",
            "c_value_estimate": "specify",
            "story": int(sheet_building.cell(column=2, row=2).value),
            "c_value": float(sheet_building.cell(column=3, row=2).value),
            "inside_pressure": sheet_building.cell(column=4, row=2).value
        }
    }

    rooms = [
        {
            "id": row[1].value,
            "name": row[2].value,
            "sub_name": row[3].value or '',
            "floor_area": float(row[4].value),
            "volume": float(row[5].value),
            "ventilation": {
                "natural": float(row[6].value)
            },
            "furniture": {
                "input_method": "default"
            },
            "schedule": {
                "name": row[7].value
            }
        } for row in sheet_rooms.iter_rows(min_row=2, max_row=n_rooms+1)
    ]

    layers_master = [
        {
            "name": row[1].value,
            "layers": make_dictionary_of_layer(row)[0],
            "reversed_layers": make_dictionary_of_layer(row)[1]
        } for row in sheet_layers.iter_rows(min_row=2, max_row=n_layers+1)
    ]

    external_general_parts = [
        {
            "id": row[1].value,
            "name": row[2].value,
            "sub_name": row[3].value or '',
            "connected_room_id": int(row[4].value),
            "boundary_type": "external_general_part",
            "area": float(row[5].value),
            "h_c": get_h_c(direction=row[8].value),  # 室内側表面対流熱伝達率, W/m2K
            "is_solar_absorbed_inside": bool(row[6].value),
            "is_floor": bool(row[6].value),
            "layers": get_layers(layers_master, layer_name=row[7].value),
            "solar_shading_part": {"existence": False},
            "is_sun_striked_outside": row[9].value == 1.0,  # fix by kitamura
            "direction": row[8].value,
            "outside_emissivity": 0.9,
            "outside_heat_transfer_resistance": get_outside_heat_transfer_resistance(direction=row[8].value, temp_dif_coef=float(row[9].value)),
            "outside_solar_absorption": 0.8,
            "temp_dif_coef": float(row[9].value)
        } for row in sheet_external_general_parts.iter_rows(min_row=2, max_row=n_external_general_parts+1)
        if float(row[5].value) != 0.0
    ]

    external_opaque_parts = [
        {
            "id": row[1].value,
            "name": row[2].value,
            "sub_name": row[3].value or '',
            "connected_room_id": int(row[4].value),
            "boundary_type": "external_opaque_part",
            "area": float(row[5].value),
            "h_c": get_h_c(direction=row[7].value),
            "is_solar_absorbed_inside": False,
            "is_floor": False,
            "solar_shading_part": {"existence": False},
            "is_sun_striked_outside": True,
            "direction": row[7].value,
            "outside_emissivity": 0.9,
            "outside_heat_transfer_resistance": get_outside_heat_transfer_resistance(direction=row[7].value, temp_dif_coef=1.0),
            "u_value": float(row[6].value),
            "inside_heat_transfer_resistance": 0.11,
            "outside_solar_absorption": 0.8,
            "temp_dif_coef": 1.0
        } for row in sheet_external_opaque_parts.iter_rows(min_row=2, max_row=n_external_opaque_parts+1)
        if float(row[5].value) != 0.0
    ]

    external_transparent_parts = [
        {
            "id": row[1].value,
            "name": row[2].value,
            "sub_name": row[3].value or '',
            "connected_room_id": int(row[4].value),
            "boundary_type": "external_transparent_part",
            "area": float(row[5].value),
            "h_c": get_h_c(direction=row[10].value),
            "is_solar_absorbed_inside": False,
            "is_floor": False,
            "solar_shading_part": get_solar_shading(exist=bool(row[11].value), depth=row[12].value, d_h=row[13].value, d_e=row[14].value),
            "is_sun_striked_outside": True,
            "direction": row[10].value,
            "outside_emissivity": 0.9,
            "outside_heat_transfer_resistance": get_outside_heat_transfer_resistance(direction=row[10].value, temp_dif_coef=1.0),
            "u_value": float(row[6].value),
            "inside_heat_transfer_resistance": 0.11,
            "eta_value": float(row[7].value),
            "incident_angle_characteristics": row[8].value,
            "glass_area_ratio": float(row[9].value),
            "temp_dif_coef": 1.0
        } for row in sheet_external_transparent_parts.iter_rows(min_row=2, max_row=n_external_transparent_parts+1)
        if float(row[5].value) != 0.0
    ]

    internals_2d = [
        [
            {
                "id": row[1].value,
                "name": row[3].value,
                "sub_name": row[5].value or '',
                "connected_room_id": int(row[7].value),
                "boundary_type": "internal",
                "area": float(row[9].value),
                "h_c": get_h_c(direction=row[11].value)[0],
                "is_solar_absorbed_inside": get_is_floor(direction=row[11].value)[0],
                "is_floor": get_is_floor(direction=row[11].value)[0],
                "layers": get_layers(layers_master, layer_name=row[10].value, is_reverse=False),
                "rear_surface_boundary_id": row[2].value
            },
            {
                "id": row[2].value,
                "name": row[4].value,
                "sub_name": row[6].value or '',
                "connected_room_id": int(row[8].value),
                "boundary_type": "internal",
                "area": float(row[9].value),
                "h_c": get_h_c(direction=row[11].value)[1],
                "is_solar_absorbed_inside": get_is_floor(direction=row[11].value)[1],
                "is_floor": get_is_floor(direction=row[11].value)[1],
                "layers": get_layers(layers_master, layer_name=row[10].value, is_reverse=True),
                "rear_surface_boundary_id": row[1].value
            }
        ] for row in sheet_internals.iter_rows(min_row=2, max_row=n_internals+1)
    ]
    # flatten
    internals = sum(internals_2d, [])

    grounds = [
        {
            "id": row[1].value,
            "name": row[2].value,
            "sub_name": row[3].value or '',
            "connected_room_id": int(row[4].value),
            "boundary_type": "ground",
            "area": float(row[5].value),
            "is_solar_absorbed_inside": bool(row[7].value),
            "is_floor": True,
            "h_c": get_h_c(direction='bottom'),
            "layers": get_layers(layers_master, layer_name=row[6].value)
        } for row in sheet_grounds.iter_rows(min_row=2, max_row=n_grounds+1)
    ]

    # NOTE: 基礎断熱における床下空間の室間換気風量をどう設定するか要確認
    ventilation_rate = 0.5
    V_MR, V_OR, V_NR = [_["volume"] for _ in rooms[:3]]
    v_vent_MR = ventilation_rate * (V_MR + V_NR * V_MR / (V_MR + V_OR))
    v_vent_OR = ventilation_rate * (V_OR + V_NR * V_OR / (V_MR + V_OR))
    mechanical_ventilations = [
        {
            "id": 0,
            "root_type": "type3",
            "volume": v_vent_MR,
            "root": [
                0,
                2
            ]
        },
        {
            "id": 1,
            "root_type": "type3",
            "volume": v_vent_OR,
            "root": [
                1,
                2
            ]
        }
    ]

    equipment_c_MR, equipment_h_MR = create_equipments(id=0, space_id=0, a_floor_is=rooms[0]['floor_area'])
    equipment_c_OR, equipment_h_OR = create_equipments(id=1, space_id=1, a_floor_is=rooms[1]['floor_area'])
    equipments = {
        "heating_equipments": [equipment_h_MR, equipment_h_OR],
        "cooling_equipments": [equipment_c_MR, equipment_c_OR]
    }

    return {
        "common": common,
        "building": building,
        "rooms": rooms,
        "boundaries": external_general_parts + external_opaque_parts + external_transparent_parts + internals + grounds,
        "mechanical_ventilations": mechanical_ventilations,
        "equipments": equipments
    }


def count_number_in_id_row(sheet):
    id_all = [row[1].value for row in sheet.rows][1:]
    return len(id_all) - (id_all).count(None)


def make_dictionary_of_layer(row):
    n = int(row[2].value)
    layer = [
        {
            "name": row[3+3*i].value,
            "thermal_resistance": float(row[4+3*i].value),
            "thermal_capacity": float(row[5+3*i].value)
        } for i in range(n)
    ]
    # Tuple(layer_list, reversed_layer_list)
    return layer, layer[::-1]


def get_layers(layers_master, layer_name, is_reverse=False):
    # use variable 'layers_master' as global variable
    layers = list(filter(lambda d: d['name'] == layer_name, layers_master))
    if len(layers) > 1:
        raise Exception("Match over one layer.")
    if len(layers) == 0:
        raise Exception("Can't find the layer")
    if is_reverse:
        return layers[0]['reversed_layers']
    else:
        return layers[0]['layers']


def get_h_c(direction):
    """室内側表面対流熱伝達率, W/m2K"""
    if direction in ['s', 'sw', 'w', 'nw', 'n', 'ne', 'e', 'se']:
        return 2.5
    elif direction == 'bottom':
        return 0.7
    elif direction == 'top':
        return 5.0
    elif direction == 'horizontal':
        return (2.5, 2.5)
    elif direction == 'upward':
        return (5.0, 0.7)
    elif direction == 'downward':
        return (0.7, 5.0)
    else:
        raise ValueError(direction)


def get_outside_heat_transfer_resistance(direction, temp_dif_coef):
    is_parting = (temp_dif_coef != 1.0)

    if direction in ['s', 'sw', 'w', 'nw', 'n', 'ne', 'e', 'se']:
        return 0.04 if not is_parting else 0.11
    elif direction == 'bottom':
        return 0.15
    elif direction == 'top':
        return 0.04 if not is_parting else 0.09
    else:
        raise Exception()


def get_solar_shading(exist: bool, depth=None, d_h=None, d_e=None):
    if exist:
        return {
            "existence": True,
            "input_method": "simple",
            "depth": float(depth),
            "d_h": float(d_h),
            "d_e": float(d_e)
        }
    else:
        return {
            "existence": False
        }


def get_is_floor(direction):
    if direction in ['s', 'sw', 'w', 'nw', 'n', 'ne', 'e', 'se', 'top']:
        return False
    elif direction == 'bottom':
        return True
    elif direction == 'horizontal':
        return (False, False)
    elif direction == 'upward':
        return (False, True)
    elif direction == 'downward':
        return (True, False)
    else:
        raise ValueError(direction)


def create_equipments(id, space_id, a_floor_is):
    q_rtd_c = 190.5 * a_floor_is + 45.6
    q_rtd_h = 1.2090 * q_rtd_c - 85.1

    q_max_c = max(0.8462 * q_rtd_c + 1205.9, q_rtd_c)
    q_max_h = max(1.7597 * q_max_c - 413.7, q_rtd_h)

    q_min_c = 500
    q_min_h = 500

    v_max_c = 11.076 * (q_rtd_c / 1000.0) ** 0.3432
    v_max_h = 11.076 * (q_rtd_h / 1000.0) ** 0.3432

    v_min_c = v_max_c * 0.55
    v_min_h = v_max_h * 0.55

    bf_c = 0.2
    bf_h = 0.2

    cooling_equipment = {
        "id": id,
        "name": f"cooling_equipment no.{id}",
        "equipment_type": "rac",
        "property": {
            "space_id": space_id,
            "q_min": q_min_c,
            "q_max": q_max_c,
            "v_min": v_min_c,
            "v_max": v_max_c,
            "bf": bf_c
        }
    }

    heating_equipment = {
        "id": id,
        "name": f"heating_equipment no.{id}",
        "equipment_type": "rac",
        "property": {
            "space_id": space_id,
            "q_min": q_min_h,
            "q_max": q_max_h,
            "v_min": v_min_h,
            "v_max": v_max_h,
            "bf": bf_h
        }
    }

    return cooling_equipment, heating_equipment


if __name__ == '__main__':
    import os
    import json
    import shutil

    dirpath = os.path.dirname(__file__)
    input_xlsx_dirpath = os.path.join(dirpath, 'input_xlsx')
    input_json_basepath = os.path.join(dirpath, 'input_json')

    # input_jsonディレクトリが残っている場合は削除する(古い入力JSONの混在を防ぐため)
    if os.path.exists(input_json_basepath):
        print(f'Remove an old directory: {input_json_basepath}')
        shutil.rmtree(input_json_basepath)

    # input_jsonディレクトリを再度作成(生成した入力JSONをこのディレクトリに入れていく)
    print(f"Make a new directory: {input_json_basepath}")
    os.makedirs(input_json_basepath)

    input_xlsx_filepaths = [_.path for _ in os.scandir(input_xlsx_dirpath) if _.is_file() and _.path.endswith('.xlsx')]
    for input_xlsx_filepath in input_xlsx_filepaths:
        input_json = convert_to_input_json(input_xlsx_filepath)

        xlsx_index, _ = os.path.splitext(os.path.basename((input_xlsx_filepath)))
        input_json_filepath = os.path.join(input_json_basepath, "house_data_{}.json".format(xlsx_index))
        print(input_json_filepath)
        with open(input_json_filepath, mode='w') as input_json_file:
            json.dump(input_json, input_json_file, indent=4)
